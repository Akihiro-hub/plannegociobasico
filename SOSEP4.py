import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side  # 必要なモジュールをインポート

# Secretsからパスワードを取得
PASSWORD = st.secrets["PASSWORD"]

# パスワード認証の処理
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if "login_attempts" not in st.session_state:
    st.session_state.login_attempts = 0

def verificar_contraseña():
    contraseña_ingresada = st.text_input("Introduce la contraseña:", type="password")

    if st.button("Iniciar sesión"):
        if st.session_state.login_attempts >= 3:
            st.error("Has superado el número máximo de intentos. Acceso bloqueado.")
        elif contraseña_ingresada == PASSWORD:  # Secretsから取得したパスワードで認証
            st.session_state.authenticated = True
            st.success("¡Autenticación exitosa! Marque otra vez el botón 'Iniciar sesión'.")
        else:
            st.session_state.login_attempts += 1
            intentos_restantes = 3 - st.session_state.login_attempts
            st.error(f"Contraseña incorrecta. Te quedan {intentos_restantes} intento(s).")
        
        if st.session_state.login_attempts >= 3:
            st.error("Acceso bloqueado. Intenta más tarde.")

if st.session_state.authenticated:
    # 認証成功後に表示されるメインコンテンツ
    st.write("## :blue[Plan de negocio en operación]") 
    st.write("###### Esta herramienta facilita la planificación del monto a vender y el flujo de caja. :green[(GuateCrece)]")  
    
    def calculate_cash_flow(initial_cash, sales, material_cost, labor_cost, loan_repayment, other_fixed_costs, desired_profit):
        fixed_cost = labor_cost + loan_repayment + other_fixed_costs
        variable_ratio = material_cost / sales
        breakeven_sales = fixed_cost / (1 - variable_ratio)
        required_sales = (fixed_cost + desired_profit) / (1 - variable_ratio)
        
        cash_flow = {
            "Saldo del efecutivo al inicio": [],
            "Ingresos (Caja de entradas)": [],
            "Egresos (Caja de salidas)": [],
            "Saldo al final": []
        }
        for month in range(12):
            cash_inflow = sales
            cash_outflow = material_cost + labor_cost + loan_repayment + other_fixed_costs
            month_end_cash = initial_cash + cash_inflow - cash_outflow
            cash_flow["Saldo del efecutivo al inicio"].append(initial_cash)
            cash_flow["Ingresos (Caja de entradas)"].append(cash_inflow)
            cash_flow["Egresos (Caja de salidas)"].append(cash_outflow)
            cash_flow["Saldo al final"].append(month_end_cash)
            initial_cash = month_end_cash
        return breakeven_sales, required_sales, cash_flow, fixed_cost, variable_ratio
    
    def generate_excel(cash_flow):
        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto del flujo de caja"
    
        headers = ["", "1r mes", "2do mes", "3r mes", "4to mes", "5to mes", "6to mes", "7mo mes", "8vo mes", "9no mes", "10mo mes", "11 mes", "12 mes"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
    
        for row_num, (key, values) in enumerate(cash_flow.items(), 2):
            ws.cell(row=row_num, column=1, value=key)
            for col_num, value in enumerate(values, 2):
                ws.cell(row=row_num, column=col_num, value=value)
    
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        return excel_data
    
    col1, col2 = st.columns(2)
    with col1:
        sales = st.number_input("Monto estimado de venta mensual (¿Cuánto monto su negocio vende al mes en GTQ?):", min_value=0, value=4100, step=1, format="%d")
        desired_profit = st.number_input("Meta de ganancias mensuales (¿Cuánto desea ganar al mes en GTQ?):", min_value=0, value=2000, step=1, format="%d")
        initial_cash = st.number_input("Saldo inicial del efecutivo (¿Cuánto monto de efecutivo comercial tiene actualmente en GTQ?):", min_value=0, value=1200, step=1, format="%d")
    with col2:
        material_cost = st.number_input("Costo mensual de materias primas (y otros costos variables, GTQ):", min_value=0, value=1500, step=1, format="%d")
        labor_cost = st.number_input("Remuneraciones mensuales de trabajadores como costo fijo (GTQ):", min_value=0, value=1200, step=1, format="%d")
        loan_repayment = st.number_input("Pago mensual de deuda (como costo fijo, GTQ):", min_value=0, value=0, step=1, format="%d")
        other_fixed_costs = st.number_input("Otros costos fijos, tales como alquiler de la tienda, electricidad, etc (GTQ):", min_value=0, value=1100, step=1, format="%d")
       
    if st.button("Elaborar el plan operativo de negocio (planificación de venta y flujo de caja)"):
        breakeven_sales, required_sales, cash_flow, fixed_cost, variable_ratio = calculate_cash_flow(
            initial_cash, sales, material_cost, labor_cost, loan_repayment, other_fixed_costs, desired_profit)
    
        st.write("#### :blue[(1) Planificación de ventas, en base al análisis del punto de equilibrio]") 
        st.write(f"Ventas al mes en el punto de equilibrio: {breakeven_sales:.2f} GTQ")
        st.write(f"Ventas necesarias para lograr la meta de ganancias al mes: {required_sales:.2f} GTQ")
    
        fig, ax = plt.subplots()
        
        sales_range = list(range(int(breakeven_sales * 0.8), int(required_sales * 1.2), 100))
        total_costs = [fixed_cost + (variable_ratio * s) for s in sales_range]
        
        ax.plot(sales_range, total_costs, color='skyblue', label="Costos totales (Costos fijos + Costos variables)", marker='o')
        ax.plot(sales_range, sales_range, color='orange', label="Venta", marker='o')
        
        ax.set_title("Análisis de punto de equilibrio")
        ax.set_xlabel("Venta (GTQ)")
        ax.set_ylabel("Costos y ventas (GTQ)")
        
        ax.axvline(breakeven_sales, color='red', linestyle='--', label=f"Punto de equilibrio: {breakeven_sales:.2f} GTQ")
        
        ax.fill_between(sales_range, total_costs, sales_range, where=[s > breakeven_sales for s in sales_range], color='skyblue', alpha=0.3, interpolate=True)
        
        mid_x = (required_sales + breakeven_sales) / 2
        mid_y = (max(total_costs) + max(sales_range)) / 2
        ax.text(mid_x, mid_y, "Ganancia = Área del color azul claro", color="blue", fontsize=7, ha="center")
    
        ax.legend()  # Show the legend
        st.pyplot(fig)
        
        months = ["1r mes", "2do mes", "3r mes", "4to mes", "5to mes", "6to mes", "7mo mes", "8vo mes", "9no mes", "10mo mes", "11 mes", "12 mes"]
        df = pd.DataFrame(cash_flow, index=months).T
        st.write("#### :blue[(2) Presupuesto del flujo de caja por 12 meses]") 
        st.dataframe(df)
    
        excel_data = generate_excel(cash_flow)
        st.download_button(
            label="Descargar la tabla EXCEL",
            data=excel_data,
            file_name="business_plan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.write("###### Puede descargar la tabla en Excel. Es recomendable elaborar el plan del flujo de caja de manera más precisa, aplicando la otra herramienta, puesto que la tabla presentada arriba es de versión muy resumida.") 

else:
    verificar_contraseña()
    
